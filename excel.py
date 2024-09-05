import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

class Excel(object):
    def __init__(self):
        return
    
    def dict_to_pandas_excel(self, dictionary: list):
        dia_atual = str(datetime.datetime.today()).split(".")[0].replace(":", "-")
        planilha_path = "./planilhas/pandas-" + str(dia_atual) + ".xlsx"
        tabela = pd.DataFrame(dictionary)
        tabela.to_excel(planilha_path, index=False)
        wb = load_workbook(planilha_path)
        ws = wb.active

        for cell in ws[1]:
            cell.font = Font(size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F65BD", end_color="2F65BD", fill_type="solid")

        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para pular o cabeçalho
            cell = ws.cell(row=row, column=2)  # Coluna B é a 2ª coluna
            cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
            cell.font = Font(color="0000FF", underline="single")  # Define a cor e sublinha o texto do link

        # Ajustar a largura das colunas, exceto a coluna B
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Letra da coluna
            if col_letter == get_column_letter(2):  # Definir largura da coluna B manualmente
                ws.column_dimensions[col_letter].width = len("Link do Leilão") + 2
                continue
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
        wb.save(planilha_path)
        wb.close()
        return planilha_path